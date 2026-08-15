"""Extract FarmNow Excel workbook structure, tables, formulas, and sample data."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\farmnow")
XLSX = ROOT / "FarmNow_ERP_System.xlsx"
OUT = ROOT / ".dist" / "excel_analysis.json"
MD = ROOT / ".dist" / "excel_analysis.md"


def cell_value(cell):
    v = cell.value
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return v


def main() -> None:
    wb = load_workbook(XLSX, data_only=False)
    wb_data = load_workbook(XLSX, data_only=True)

    result = {
        "file": str(XLSX.name),
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "defined_names": {},
        "sheets": {},
    }

    for name, dn in wb.defined_names.items():
        try:
            result["defined_names"][name] = {
                "attr_text": dn.attr_text,
                "hidden": getattr(dn, "hidden", None),
            }
        except Exception as e:
            result["defined_names"][name] = {"error": str(e)}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        sheet_info = {
            "title": ws.title,
            "sheet_state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimensions": ws.dimensions,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "tables": {},
            "data_validations": [],
            "formulas": [],
            "sample_rows": [],
            "header_row": None,
            "charts": [],
            "print_area": str(ws.print_area) if ws.print_area else None,
        }

        # Charts
        for i, chart in enumerate(getattr(ws, "_charts", []) or []):
            sheet_info["charts"].append(
                {
                    "index": i,
                    "type": type(chart).__name__,
                    "title": str(getattr(chart, "title", None)),
                    "anchor": str(getattr(chart, "anchor", None)),
                }
            )

        # Tables — TableList.items() yields name -> ref string; values() yields Table objects
        for table in ws.tables.values():
            cols = [c.name for c in (table.tableColumns or [])]
            sheet_info["tables"][table.displayName] = {
                "displayName": table.displayName,
                "ref": table.ref,
                "headerRowCount": table.headerRowCount,
                "totalsRowCount": table.totalsRowCount,
                "columns": cols,
            }

        # Data validations
        dvs = ws.data_validations.dataValidation if ws.data_validations else []
        for dv in dvs:
            sheet_info["data_validations"].append(
                {
                    "sqref": str(dv.sqref),
                    "type": dv.type,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": dv.allow_blank,
                    "showErrorMessage": dv.showErrorMessage,
                    "errorTitle": dv.errorTitle,
                    "error": dv.error,
                }
            )

        # Scan used range for formulas and values
        max_r = min(ws.max_row or 0, 200)
        max_c = min(ws.max_column or 0, 40)
        headers = []
        for c in range(1, max_c + 1):
            headers.append(cell_value(ws.cell(1, c)))
        sheet_info["header_row"] = headers

        # Capture data rows as cached values
        rows = []
        for r in range(1, min((ws.max_row or 0) + 1, 80)):
            row = []
            for c in range(1, max_c + 1):
                row.append(cell_value(ws_data.cell(r, c)))
            if any(x is not None and x != "" for x in row):
                rows.append(row)
        sheet_info["sample_rows"] = rows

        formula_count_by_col = defaultdict(int)
        unique_formulas = {}
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    col_letter = get_column_letter(c)
                    formula_count_by_col[col_letter] += 1
                    key = (c, v)
                    if v not in unique_formulas:
                        unique_formulas[v] = {
                            "formula": v,
                            "first_cell": f"{col_letter}{r}",
                            "column_header": headers[c - 1] if c - 1 < len(headers) else None,
                        }
        sheet_info["unique_formulas"] = list(unique_formulas.values())
        sheet_info["formula_count_by_col"] = dict(formula_count_by_col)

        # Capture a wider dump of non-empty cells for dashboard / calc / report sheets
        if sheet_name.startswith(("Dashboard", "calc_", "rpt_", "Menu", "mst_Settings", "mst_Lists")):
            cells = []
            mr = min(ws.max_row or 0, 120)
            mc = min(ws.max_column or 0, 20)
            for r in range(1, mr + 1):
                for c in range(1, mc + 1):
                    fcell = ws.cell(r, c)
                    dcell = ws_data.cell(r, c)
                    if fcell.value is None and dcell.value is None:
                        continue
                    cells.append(
                        {
                            "addr": f"{get_column_letter(c)}{r}",
                            "formula": cell_value(fcell) if isinstance(fcell.value, str) and str(fcell.value).startswith("=") else None,
                            "value": cell_value(dcell),
                            "raw": cell_value(fcell) if not (isinstance(fcell.value, str) and str(fcell.value).startswith("=")) else None,
                        }
                    )
            sheet_info["notable_cells"] = cells

        result["sheets"][sheet_name] = sheet_info

    OUT.write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")

    # Markdown summary
    lines = ["# Excel analysis dump", "", f"Sheets: {len(wb.sheetnames)}", ""]
    for name in wb.sheetnames:
        s = result["sheets"][name]
        lines.append(f"## {name}")
        lines.append(f"- state: `{s['sheet_state']}`")
        lines.append(f"- size: {s['max_row']} rows x {s['max_column']} cols")
        if s["tables"]:
            for tname, t in s["tables"].items():
                lines.append(f"- table `{tname}` ref `{t['ref']}` cols: {t['columns']}")
        if s["unique_formulas"]:
            lines.append("- unique formulas:")
            for f in s["unique_formulas"][:40]:
                lines.append(f"  - `{f['first_cell']}` ({f['column_header']}): `{f['formula']}`")
        if s["data_validations"]:
            lines.append("- validations:")
            for dv in s["data_validations"]:
                lines.append(f"  - {dv['sqref']}: type={dv['type']} f1={dv['formula1']}")
        if s["charts"]:
            lines.append(f"- charts: {s['charts']}")
        lines.append("")
        if s["sample_rows"]:
            lines.append("Sample:")
            for row in s["sample_rows"][:8]:
                lines.append("  " + " | ".join("" if x is None else str(x) for x in row[:18]))
            lines.append("")
    MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"Wrote {MD}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
