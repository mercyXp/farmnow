"""Dump every Excel table's data rows to JSON for seed generation."""
from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"C:\farmnow\FarmNow_ERP_System.xlsx")
OUT = Path(r"C:\farmnow\.dist\excel_tables.json")


def serialize(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat() if v.time() == datetime.min.time() else v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    tables = {}
    for ws in wb.worksheets:
        for table in ws.tables.values():
            cols = [c.name for c in table.tableColumns]
            ref = table.ref
            # parse range
            from openpyxl.utils import range_boundaries

            min_col, min_row, max_col, max_row = range_boundaries(ref)
            rows = []
            for r in range(min_row + 1, max_row + 1):
                row = {}
                empty = True
                for i, c in enumerate(range(min_col, max_col + 1)):
                    val = serialize(ws.cell(r, c).value)
                    row[cols[i]] = val
                    if val not in (None, ""):
                        empty = False
                if not empty:
                    rows.append(row)
            tables[table.displayName] = {"sheet": ws.title, "columns": cols, "rows": rows}
    OUT.write_text(json.dumps(tables, indent=2), encoding="utf-8")
    print("tables", list(tables))
    for name, t in tables.items():
        print(f"  {name}: {len(t['rows'])} rows")


if __name__ == "__main__":
    main()
